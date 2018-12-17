from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl.cell import Cell

TITLE_COLUMN = 5
ID_COLUMN = 1
STATUS_COLUMN = 3
VERSION_COLUMN = 8


def find_in_file(file_name, text):

    wb = load_workbook(file_name)
    # 只支持第一sheet
    ws = wb.active
    for row_no in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_no, column=TITLE_COLUMN)
        if cell.data_type == Cell.TYPE_STRING and cell.value.find(text) > 0:
            return [
                ws.cell(row=cell.row, column=ID_COLUMN).value,
                ws.cell(row=cell.row, column=STATUS_COLUMN).value,
                ws.cell(row=cell.row, column=VERSION_COLUMN).value
            ]


collection_book = load_workbook('collection.xlsx')

collection_sheet = collection_book.active

for row_num in range(1, collection_sheet.max_row + 1):
    substring = collection_sheet.cell(row=row_num, column=9).value[-10:-3]
    result = find_in_file("issues.xlsx", substring)
    if result is not None:
        collection_sheet.cell(row=row_num, column=20).value = result[0]
        collection_sheet.cell(row=row_num, column=20).hyperlink =\
            "http://10.199.129.57/redmine/issues/{}".format(result[0])
        collection_sheet.cell(row=row_num, column=21).value = result[1]
        collection_sheet.cell(row=row_num, column=22).value = result[2]

collection_book.save('collection.xlsx')
