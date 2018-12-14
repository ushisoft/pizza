from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl.cell import Cell

wb = load_workbook('issues.xlsx')
# 只支持第一sheet
ws = wb.active
print(ws.max_row)
for row_no in range(1, ws.max_row):
    print(row_no)
